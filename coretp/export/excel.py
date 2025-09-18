# SPDX-FileCopyrightText: © 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

import argparse
from pathlib import Path
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .base import Formatter, ExportContext
from .formatters import TestPlanFormatter, ExportTableBuilder
from coretp.models import TestPlan


class XlsFormatter(Formatter):
    """
    Formatter for Excel files using openpyxl.
    """

    description = "Generate Excel documentation from test plans"
    help_message = "Export test plans to Excel format with multiple worksheets and formatting"
    suffix = "xlsx"

    def __init__(self, no_summary: bool = False, single_sheet: bool = False, **kwargs):
        super().__init__(**kwargs)
        self.no_summary = no_summary
        self.single_sheet = single_sheet

        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.wrap_text = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def add_arguments(parser: argparse.ArgumentParser):
        """Add Excel-specific arguments."""
        parser.add_argument("--no-summary", action="store_true", help="Skip summary worksheet")
        parser.add_argument("--single-sheet", action="store_true", help="Put all test plans in single worksheet")

    def build(self) -> Path:
        """Generate Excel file containing all test plans."""
        if not self.test_plans:
            raise ValueError("No test plans added. Use add_test_plan() first.")

        context = ExportContext(self.output_file.with_suffix(""), total_plans=len(self.test_plans), **self.metadata)

        with context.binary_file(f".{self.suffix}") as (f, output_path):
            wb = Workbook()

            # remove default sheet and create summary sheet if requested
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            if not self.no_summary:
                self._create_summary_sheet(wb)

            if not self.single_sheet:
                for i, test_plan in enumerate(self.test_plans):
                    sheet_name = self._sanitize_sheet_name(test_plan.name, i)
                    ws = wb.create_sheet(title=sheet_name)
                    self._populate_test_plan_sheet(ws, test_plan)
            else:
                ws = wb.create_sheet(title="All Test Plans")
                self._populate_combined_sheet(ws)

            wb.save(str(output_path))

        return output_path

    def _create_summary_sheet(self, wb: Workbook):
        """Create overview summary sheet."""
        ws = wb.create_sheet(title="Summary", index=0)

        # Title
        ws["A1"] = "Test Plan Documentation Summary"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells("A1:E1")

        # Metadata
        row = 3
        ws[f"A{row}"] = "Generated:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        ws[f"A{row}"] = "Total Test Plans:"
        ws[f"B{row}"] = len(self.test_plans)
        row += 1

        total_scenarios = sum(len(tp.scenarios) for tp in self.test_plans)
        ws[f"A{row}"] = "Total Scenarios:"
        ws[f"B{row}"] = total_scenarios
        row += 2

        # Test plans table using reusable builder
        table_data = ExportTableBuilder.build_summary_table(self.test_plans)

        # Add sheet reference column
        table_data[0].append("Sheet")  # Add header
        for i, test_plan in enumerate(self.test_plans):
            sheet_ref = self._sanitize_sheet_name(test_plan.name, i) if not self.single_sheet else "All Test Plans"
            table_data[i + 1].append(sheet_ref)

        # Write table
        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=row + row_idx, column=col_idx + 1, value=cell_value)
                cell.border = self.border
                if row_idx == 0:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_align

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _populate_test_plan_sheet(self, ws: Worksheet, test_plan: TestPlan):
        """Populate worksheet with single test plan data."""
        row = 1

        # Test plan header
        ws[f"A{row}"] = test_plan.name
        ws[f"A{row}"].font = Font(size=14, bold=True, color="366092")
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        ws[f"A{row}"] = test_plan.description
        ws[f"A{row}"].alignment = self.wrap_text
        ws.merge_cells(f"A{row}:F{row}")
        row += 2

        # Scenarios table using reusable builder
        table_data = ExportTableBuilder.build_scenario_table(test_plan)

        # Write table
        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=row + row_idx, column=col_idx + 1, value=cell_value)
                cell.border = self.border
                if row_idx == 0:  # Header row
                    cell.font = self.subheader_font
                    cell.fill = self.subheader_fill
                    cell.alignment = self.center_align
                else:
                    if col_idx == 1:  # Description column
                        cell.alignment = self.wrap_text

        self._auto_adjust_columns(ws)

    def _populate_combined_sheet(self, ws: Worksheet):
        """Populate worksheet with all test plans in single sheet."""
        row = 1

        # Overall header
        ws[f"A{row}"] = "All Test Plans"
        ws[f"A{row}"].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells(f"A{row}:G{row}")
        row += 2

        # Combined table headers
        headers = ["Test Plan", "Scenario", "Description", "Environment", "Steps", "Paging Modes", "Privilege Mode"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # Combined table using reusable builder
        table_data = ExportTableBuilder.build_combined_table(self.test_plans)

        # Write table (skip headers since they're already written)
        for row_idx, row_data in enumerate(table_data[1:], 1):  # Skip header row
            for col_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=row + row_idx, column=col_idx + 1, value=cell_value)
                cell.border = self.border
                if col_idx == 2:  # Description column
                    cell.alignment = self.wrap_text

        self._auto_adjust_columns(ws)

    def _sanitize_sheet_name(self, name: str, index: int) -> str:
        """Sanitize sheet name for Excel compatibility."""
        # Excel sheet names can't exceed 31 chars and can't contain certain chars
        invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")

        if len(sanitized) > 28:  # Leave room for index
            sanitized = sanitized[:28]

        return f"{sanitized}_{index+1}"

    def _auto_adjust_columns(self, ws: Worksheet):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width
